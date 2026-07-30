"""Deliberately messy sample workbooks.

Spec 3.1 lists what real SME sheets do: merged headers, mid table total rows,
dates stored as text, product names spelled three ways, currency symbols inside
numeric cells, one tab per month. Each generator here reproduces one of those,
so the detection layer is tested against the mess it exists to absorb rather
than against tidy data it would never see.

Everything is seeded, so the fixtures are identical run to run.
"""

from __future__ import annotations

from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

PRODUCTS = [
    "Linen Candle",
    "Ceramic Diffuser",
    "Soap Set",
    "Room Spray",
    "Gift Box",
]
PRICES = {
    "Linen Candle": 6500.0,
    "Ceramic Diffuser": 9800.0,
    "Soap Set": 3200.0,
    "Room Spray": 1900.0,
    "Gift Box": 4500.0,
}
CHANNELS = ["online", "in store", "wholesale"]
REGIONS = ["Lagos", "Abuja", "Kano", "Port Harcourt"]

#: The same product typed three different ways, which is the norm in a sheet
#: maintained by more than one person.
PRODUCT_ALIASES = {
    "Linen Candle": ["Linen Candle", "linen candle", "Linen  Candle"],
    "Ceramic Diffuser": ["Ceramic Diffuser", "Ceramic diffuser", "CERAMIC DIFFUSER"],
}


def base_sales(n: int = 420, seed: int = 42) -> pd.DataFrame:
    """A clean, well formed sales table. The control case."""
    rng = np.random.default_rng(seed)
    dates = pd.date_range("2025-01-01", periods=n, freq="D")
    products = rng.choice(PRODUCTS, n)
    quantity = rng.integers(1, 6, n)
    unit_price = np.array([PRICES[p] for p in products])
    revenue = quantity * unit_price
    return pd.DataFrame(
        {
            "order_date": dates,
            "product_name": products,
            "quantity": quantity,
            "unit_price": unit_price,
            "total_paid": revenue,
            "unit_cost": np.round(unit_price * 0.55, 2),
            "customer_id": rng.integers(1000, 1150, n),
            "channel": rng.choice(CHANNELS, n),
        }
    )


def messy_frame(n: int = 420, seed: int = 7) -> pd.DataFrame:
    """Every content-level sin at once, still as a plain DataFrame."""
    rng = np.random.default_rng(seed)
    clean = base_sales(n, seed)

    def spell(product: str) -> str:
        options = PRODUCT_ALIASES.get(product)
        return rng.choice(options) if options else product

    return pd.DataFrame(
        {
            # Badly named but perfectly parseable dates, stored as text.
            "Column3": clean["order_date"].dt.strftime("%d/%m/%Y"),
            # Named like a date, holds a place. The spec's own veto example.
            "date": rng.choice(REGIONS, n),
            "Item Descrption": [spell(p) for p in clean["product_name"]],
            "qty": clean["quantity"],
            # Currency symbols and thousands separators inside numeric cells.
            "Amount (NGN)": [f"N{v:,.2f}" for v in clean["total_paid"]],
            # The trap: must not be read as revenue.
            "discount_amount": np.where(
                rng.random(n) < 0.8, 0.0, np.round(clean["total_paid"] * 0.1, 2)
            ),
            "unit_cst": clean["unit_cost"],
            "src": clean["channel"],
            # Unknown but usable as a grouping.
            "salesperson": rng.choice(["Ada", "Bola", "Chidi"], n),
            # Entirely empty spacer column.
            "notes": [""] * n,
        }
    )


def write_messy_workbook(path: str | Path, n: int = 240, seed: int = 5) -> Path:
    """A workbook with a title banner, a merged two-row header and total rows.

    This is the structural mess, as opposed to the content mess above. It has
    to be written with openpyxl because pandas cannot produce a genuinely
    merged header cell.
    """
    path = Path(path)
    rng = np.random.default_rng(seed)
    wb = Workbook()
    ws = wb.active
    ws.title = "Sales"

    # Rows 1-2: a title banner nobody asked for.
    ws["A1"] = "FERN & FLAME LTD"
    ws["A2"] = "Sales Report — January to June 2025"

    # Rows 4-5: a merged two-row header.
    ws.merge_cells(start_row=4, start_column=3, end_row=4, end_column=4)
    ws.cell(row=4, column=3, value="Order")
    ws.merge_cells(start_row=4, start_column=5, end_row=4, end_column=6)
    ws.cell(row=4, column=5, value="Money")

    headers = ["Date", "Product", "Qty", "Price", "Total", "Cost"]
    for i, name in enumerate(headers, start=1):
        ws.cell(row=5, column=i, value=name)

    row = 6
    written = 0
    start = pd.Timestamp("2025-01-01")
    for i in range(n):
        product = str(rng.choice(PRODUCTS))
        qty = int(rng.integers(1, 6))
        price = PRICES[product]
        when = start + pd.Timedelta(days=i)

        ws.cell(row=row, column=1, value=when.strftime("%d/%m/%Y"))
        ws.cell(row=row, column=2, value=product)
        ws.cell(row=row, column=3, value=qty)
        ws.cell(row=row, column=4, value=price)
        ws.cell(row=row, column=5, value=qty * price)
        ws.cell(row=row, column=6, value=round(price * 0.55, 2))
        row += 1
        written += 1

        # A subtotal row every 30 transactions, and a blank spacer after it.
        if written % 30 == 0:
            ws.cell(row=row, column=1, value="Subtotal")
            ws.cell(row=row, column=5, value=999999)
            row += 2

    ws.cell(row=row, column=1, value="GRAND TOTAL")
    ws.cell(row=row, column=5, value=123456789)

    wb.save(path)
    return path


def write_monthly_workbook(path: str | Path, months: int = 6, seed: int = 9) -> Path:
    """One tab per month, identical layout. Should combine into one table."""
    path = Path(path)
    rng = np.random.default_rng(seed)
    wb = Workbook()
    wb.remove(wb.active)

    for m in range(1, months + 1):
        ws = wb.create_sheet(title=pd.Timestamp(2025, m, 1).strftime("%b"))
        for i, name in enumerate(
            ["order_date", "product", "qty", "unit_price", "revenue"], start=1
        ):
            ws.cell(row=1, column=i, value=name)
        days = pd.Period(f"2025-{m:02d}").days_in_month
        for d in range(days):
            product = str(rng.choice(PRODUCTS))
            qty = int(rng.integers(1, 6))
            price = PRICES[product]
            ws.cell(row=d + 2, column=1, value=f"2025-{m:02d}-{d + 1:02d}")
            ws.cell(row=d + 2, column=2, value=product)
            ws.cell(row=d + 2, column=3, value=qty)
            ws.cell(row=d + 2, column=4, value=price)
            ws.cell(row=d + 2, column=5, value=qty * price)

    wb.save(path)
    return path


def planted_business(seed: int = 21, days: int = 540) -> pd.DataFrame:
    """A business with deliberately planted, known truths.

    Every effect below is put there on purpose so the analysis engine can be
    checked against ground truth rather than against whatever it happens to
    say. The planted facts are:

    * a real decline from month 7 onward, driven by the online channel
    * Linen Candle sells most but Ceramic Diffuser earns most (margin reality)
    * profit is concentrated in Ceramic Diffuser
    * Gift Box sells at a genuine loss
    * no relationship at all between salesperson and order value
    """
    rng = np.random.default_rng(seed)
    rows: list[dict] = []

    # Cost ratios chosen so the best seller is not the best earner.
    cost_ratio = {
        "Linen Candle": 0.78,  # high volume, thin margin
        "Ceramic Diffuser": 0.42,  # lower volume, fat margin
        "Soap Set": 0.60,
        "Room Spray": 0.65,
        "Gift Box": 1.18,  # sold at a loss, on purpose
    }
    weights = [0.40, 0.22, 0.16, 0.14, 0.08]

    start = pd.Timestamp("2024-01-01")
    for day in range(days):
        when = start + pd.Timedelta(days=day)
        month_index = day // 30

        # The planted decline: online falls away after month 7.
        decline = 1.0 if month_index < 7 else max(0.45, 1.0 - 0.06 * (month_index - 6))
        orders = rng.poisson(6)

        for _ in range(orders):
            product = str(rng.choice(PRODUCTS, p=weights))
            channel = str(rng.choice(CHANNELS, p=[0.55, 0.30, 0.15]))
            factor = decline if channel == "online" else 1.0
            if rng.random() > factor:
                continue  # the lost online orders

            qty = int(rng.integers(1, 5))
            price = PRICES[product] * float(rng.normal(1.0, 0.04))
            rows.append(
                {
                    "order_date": when,
                    "product_name": product,
                    "quantity": qty,
                    "unit_price": round(price, 2),
                    "total_paid": round(price * qty, 2),
                    "unit_cost": round(PRICES[product] * cost_ratio[product], 2),
                    "customer_id": int(rng.integers(1000, 1400)),
                    "channel": channel,
                    # Deliberately unrelated to anything: must not produce a
                    # finding once multiple comparisons are accounted for.
                    "salesperson": str(rng.choice(["Ada", "Bola", "Chidi", "Dele"])),
                }
            )

    return pd.DataFrame(rows)


def flat_business(seed: int = 33, days: int = 540) -> pd.DataFrame:
    """A business with no trend, no concentration and no real differences.

    The engine must not manufacture findings from this. It is the control that
    catches an engine tuned to always have something to say.
    """
    rng = np.random.default_rng(seed)
    rows: list[dict] = []
    start = pd.Timestamp("2024-01-01")
    for day in range(days):
        when = start + pd.Timedelta(days=day)
        for _ in range(rng.poisson(6)):
            product = str(rng.choice(PRODUCTS))  # uniform: no concentration
            qty = int(rng.integers(1, 5))
            price = 5000.0 * float(rng.normal(1.0, 0.05))
            rows.append(
                {
                    "order_date": when,
                    "product_name": product,
                    "quantity": qty,
                    "unit_price": round(price, 2),
                    "total_paid": round(price * qty, 2),
                    "unit_cost": round(5000.0 * 0.6, 2),
                    "channel": str(rng.choice(CHANNELS)),
                    "salesperson": str(rng.choice(["Ada", "Bola", "Chidi", "Dele"])),
                }
            )
    return pd.DataFrame(rows)


def seasonal_business(seed: int = 44, years: int = 3) -> pd.DataFrame:
    """A flat business with a strong, repeating Christmas peak and January dip.

    Underlying demand does not change at all across the years. Only the
    calendar moves. The engine must therefore *not* report the January fall as
    a decline, which is the whole point of deseasonalising (spec 5).
    """
    rng = np.random.default_rng(seed)
    # Index by calendar month: a big December, a weak January.
    shape = {
        1: 0.55, 2: 0.75, 3: 0.90, 4: 0.95, 5: 1.00, 6: 1.00,
        7: 0.95, 8: 0.95, 9: 1.05, 10: 1.20, 11: 1.60, 12: 2.10,
    }
    rows: list[dict] = []
    start = pd.Timestamp("2023-01-01")
    for day in range(365 * years):
        when = start + pd.Timedelta(days=day)
        expected = 6 * shape[when.month]
        for _ in range(rng.poisson(expected)):
            product = str(rng.choice(PRODUCTS))
            qty = int(rng.integers(1, 5))
            price = PRICES[product] * float(rng.normal(1.0, 0.04))
            rows.append(
                {
                    "order_date": when,
                    "product_name": product,
                    "quantity": qty,
                    "unit_price": round(price, 2),
                    "total_paid": round(price * qty, 2),
                }
            )
    return pd.DataFrame(rows)


def write_clean_workbook(path: str | Path, n: int = 300, seed: int = 42) -> Path:
    """A tidy single-sheet workbook. Must reach analysis with zero questions."""
    path = Path(path)
    base_sales(n, seed).to_excel(path, index=False)
    return path
